"""
Liest TAXI-Lohnlisten (Sheet1, Header in Zeile 4) und extrahiert
pro Mitarbeiter die Monatswerte plus Stundensatz aus Excel-Spalte S.

Unterschiede zum Mietwagen-Parser:
- Header nicht in Zeile 1, sondern in Zeile 4
- Daten ab Zeile 6 (Zeile 5 ist oft leer)
- PersNr-Spalte (B) statt aus Mitarbeiter-Tab-Namen
- Keine eigenen Mitarbeiter-Sheets — alles im Sheet1
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl

from mapping import (
    DATEN_START_ROW,
    HEADER_ROW,
    INFO_COL,
    KONTROLL_BRUTTO_COL,
    LOHNART_MAPPING,
    MANUELL_IN_DATEV,
    STUNDENSATZ_COL,
)

NAME_COL = "A"
PERSNR_COL = "B"


@dataclass
class MitarbeiterZeile:
    name: str
    pers_nr: Optional[str]
    stundensatz: float = 0.0
    werte: Dict[str, float] = field(default_factory=dict)
    manuell_werte: Dict[str, float] = field(default_factory=dict)
    soll_brutto: float = 0.0
    info: Optional[str] = None
    warnungen: List[str] = field(default_factory=list)


@dataclass
class ParseResult:
    mitarbeiter: List[MitarbeiterZeile]
    globale_warnungen: List[str] = field(default_factory=list)


def _col_letter_to_index(letter: str) -> int:
    return openpyxl.utils.column_index_from_string(letter)


def _zahl(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in ("######", "#####", "#ZAHL!", "#WERT!"):
        return 0.0
    # "85,48 h" oder "1.265,08" → 85.48 / 1265.08
    s = s.replace(" h", "").replace("h", "")
    # Tausendertrennzeichen und Dezimal-Komma
    if "," in s and "." in s:
        # z.B. "1.265,08" → "1265.08"
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_header(value) -> str:
    s = "" if value is None else str(value)
    s = s.replace("€", "").replace(".", "").replace(",", "")
    s = " ".join(s.split())
    return s.strip().casefold()


def _validate_headers(ws) -> List[str]:
    """Prüft Header-Zeile gegen erwartete Spalten."""
    erwartet: Dict[str, str] = {}
    for m in LOHNART_MAPPING:
        erwartet[m["excel_col"]] = m["excel_header"]
    for u in MANUELL_IN_DATEV:
        erwartet[u["excel_col"]] = u["excel_header"]
    erwartet[STUNDENSATZ_COL] = "Stundensatz"
    erwartet[KONTROLL_BRUTTO_COL] = "Brutto"

    probleme: List[str] = []
    for col, exp in erwartet.items():
        idx = _col_letter_to_index(col)
        tatsaechlich = ws.cell(HEADER_ROW, idx).value
        if _norm_header(tatsaechlich) != _norm_header(exp):
            probleme.append(
                f'Spalte {col}: erwartet "{exp}", gefunden "{tatsaechlich}". '
                f'Wenn das Schema bei diesem Mandant anders ist, mapping.py anpassen.'
            )
    return probleme


def parse_excel(file_bytes: bytes) -> ParseResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active  # Taxi-Excels haben typischerweise nur "Sheet1"

    header_probleme = _validate_headers(ws)
    if header_probleme:
        return ParseResult(
            mitarbeiter=[],
            globale_warnungen=[
                "Excel-Schema weicht von erwartetem Aufbau ab — Abbruch, um Datenverwechslung zu verhindern.",
                *header_probleme,
            ],
        )

    name_idx = _col_letter_to_index(NAME_COL)
    persnr_idx = _col_letter_to_index(PERSNR_COL)
    info_idx = _col_letter_to_index(INFO_COL)
    stundensatz_idx = _col_letter_to_index(STUNDENSATZ_COL)
    brutto_idx = _col_letter_to_index(KONTROLL_BRUTTO_COL)

    mapped_cols = [(m, _col_letter_to_index(m["excel_col"])) for m in LOHNART_MAPPING]
    manuell_cols = [(u, _col_letter_to_index(u["excel_col"])) for u in MANUELL_IN_DATEV]

    mitarbeiter: List[MitarbeiterZeile] = []

    for r in range(DATEN_START_ROW, ws.max_row + 1):
        name = ws.cell(r, name_idx).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        persnr_val = ws.cell(r, persnr_idx).value
        pers_nr = str(int(persnr_val)) if isinstance(persnr_val, (int, float)) else (str(persnr_val).strip() if persnr_val else None)

        info_val = ws.cell(r, info_idx).value
        info = str(info_val).strip() if info_val else None

        zeile = MitarbeiterZeile(name=name, pers_nr=pers_nr, info=info)
        zeile.stundensatz = _zahl(ws.cell(r, stundensatz_idx).value)

        for m, col_idx in mapped_cols:
            val = _zahl(ws.cell(r, col_idx).value)
            if val == 0:
                continue
            if m.get("vorzeichen_umkehren"):
                val = -abs(val)  # immer negativ (für Vorschuss / Abschlag)
            zeile.werte[m["lohnart"]] = round(val, 2)

        for u, col_idx in manuell_cols:
            val = _zahl(ws.cell(r, col_idx).value)
            if val != 0:
                zeile.manuell_werte[u["excel_header"]] = val

        zeile.soll_brutto = _zahl(ws.cell(r, brutto_idx).value)

        # Plausibilitäts-Check: Summe Lohnarten ohne Vorschuss/Abschlag/Urlaub
        # sollte ~ Brutto sein. Toleranz 50 € — Brutto-Berechnung variiert je Firma,
        # daher konservative Schwelle.
        importable_summe = sum(
            v for la, v in zeile.werte.items()
            if la not in ("9000", "9001", "1600")  # Vorschuss/Abschlag/Urlaub nicht im Brutto
        )
        if zeile.soll_brutto > 0:
            diff = abs(importable_summe - zeile.soll_brutto)
            if diff > 50.0:
                zeile.warnungen.append(
                    f"Plausibilität: Importwerte-Summe {importable_summe:.2f} € weicht "
                    f"deutlich von Brutto {zeile.soll_brutto:.2f} € ab ({diff:.2f} €). "
                    f"Bitte Excel prüfen."
                )

        mitarbeiter.append(zeile)

    return ParseResult(mitarbeiter=mitarbeiter)


def monat_jahr_aus_dateiname(filename: str) -> Optional[Tuple[int, int]]:
    """'Lohnliste 04-26 GOS.xlsx' -> (2026, 4)."""
    m = re.search(r"(\d{1,2})[-_](\d{2,4})", filename)
    if m:
        monat, jahr = int(m.group(1)), int(m.group(2))
        if jahr < 100:
            jahr += 2000
        if 1 <= monat <= 12 and 2000 <= jahr <= 2100:
            return (jahr, monat)
    # Fallback: YYYY_MM oder YYYY-MM
    m = re.search(r"(\d{4})[-_](\d{1,2})", filename)
    if m:
        jahr, monat = int(m.group(1)), int(m.group(2))
        if 1 <= monat <= 12:
            return (jahr, monat)
    return None


def firma_aus_dateiname(filename: str) -> str:
    """'Lohnliste 04-26 GOS.xlsx' -> 'GOS'."""
    base = re.sub(r"\.xlsx?$", "", filename, flags=re.IGNORECASE)
    # Nach dem Datum kommt der Name
    m = re.search(r"\d{1,2}[-_]\d{2,4}\s+(.+)$", base)
    if m:
        return m.group(1).strip()
    m = re.search(r"\d{4}[-_]\d{1,2}\s+(.+)$", base)
    if m:
        return m.group(1).strip()
    return base.replace("Lohnliste", "").strip()
