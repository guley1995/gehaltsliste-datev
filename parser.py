"""
Liest die Mietwagen-Gehaltsliste-Excel und extrahiert pro Mitarbeiter
die Monatswerte aus dem 'Gehalt'-Sheet sowie die Personalnummer aus
dem zugehörigen Mitarbeiter-Tabnamen.
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import openpyxl

from mapping import (
    KONTROLL_GRUNDGEHALT_COL,
    LOHNART_MAPPING,
    MANUELL_IN_DATEV,
    STUNDENSATZ_COL,
)

UEBERSICHT_SHEET = "Gehalt"
NAME_COL = "A"


@dataclass
class MitarbeiterZeile:
    name: str
    pers_nr: Optional[str]
    stundensatz: float = 0.0
    werte: Dict[str, float] = field(default_factory=dict)
    manuell_werte: Dict[str, float] = field(default_factory=dict)
    soll_grundgehalt: float = 0.0
    info: Optional[str] = None
    warnungen: List[str] = field(default_factory=list)


@dataclass
class ParseResult:
    mitarbeiter: List[MitarbeiterZeile]
    globale_warnungen: List[str] = field(default_factory=list)


def _col_letter_to_index(letter: str) -> int:
    return openpyxl.utils.column_index_from_string(letter)


def _persnr_aus_tabname(tab: str) -> Optional[str]:
    m = re.search(r"_([\d.]+)\s*$", tab)
    return m.group(1) if m else None


def _baue_namen_index(sheet_names: List[str]) -> Dict[str, str]:
    """Map exakter Mitarbeitername -> Tabname. Übersichts-Sheet ausgenommen."""
    out: Dict[str, str] = {}
    for tab in sheet_names:
        if tab == UEBERSICHT_SHEET:
            continue
        prefix = re.sub(r"_[\d.]+\s*$", "", tab).strip()
        out[prefix] = tab
    return out


def _zahl(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def _norm_header(value) -> str:
    """Normalisiert einen Excel-Header für tolerante Vergleichbarkeit."""
    s = "" if value is None else str(value)
    s = s.replace("€", "").replace(".", "").replace(",", "")
    s = " ".join(s.split())  # Whitespace zusammenfassen
    return s.strip().casefold()


def _validate_headers(ws) -> List[str]:
    """Prüft, ob in Zeile 1 die erwarteten Header an den erwarteten Spalten
    stehen. Liefert Liste von Problem-Strings (leer = alles OK)."""
    erwartet: Dict[str, str] = {}
    for m in LOHNART_MAPPING:
        erwartet[m["excel_col"]] = m["excel_header"]
    for u in MANUELL_IN_DATEV:
        erwartet[u["excel_col"]] = u["excel_header"]

    probleme: List[str] = []
    for col, exp in erwartet.items():
        idx = _col_letter_to_index(col)
        tatsaechlich = ws.cell(1, idx).value
        if _norm_header(tatsaechlich) != _norm_header(exp):
            probleme.append(
                f'Spalte {col}: erwartet "{exp}", gefunden "{tatsaechlich}". '
                f"Wenn das Schema bei diesem Mandant anders ist, mapping.py anpassen."
            )
    return probleme


def parse_excel(file_bytes: bytes) -> ParseResult:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    if UEBERSICHT_SHEET not in wb.sheetnames:
        return ParseResult(
            mitarbeiter=[],
            globale_warnungen=[f"Sheet '{UEBERSICHT_SHEET}' nicht gefunden. Vorhandene Sheets: {wb.sheetnames}"],
        )

    ws = wb[UEBERSICHT_SHEET]

    # Schema-Validierung: stoppt mit Fehler, wenn Header nicht passen
    header_probleme = _validate_headers(ws)
    if header_probleme:
        return ParseResult(
            mitarbeiter=[],
            globale_warnungen=[
                "Excel-Schema weicht von erwartetem Aufbau ab — Abbruch, um Datenverwechslung zu verhindern.",
                *header_probleme,
            ],
        )

    name_to_tab = _baue_namen_index(wb.sheetnames)

    name_col_idx = _col_letter_to_index(NAME_COL)
    info_col_idx = _col_letter_to_index("V")  # Spalte INFO (war W, nach Entfernung der Flughafengebühr-Spalte O)

    mapped_cols = [(m, _col_letter_to_index(m["excel_col"])) for m in LOHNART_MAPPING]
    manuell_cols = [(u, _col_letter_to_index(u["excel_col"])) for u in MANUELL_IN_DATEV]
    kontroll_col_idx = _col_letter_to_index(KONTROLL_GRUNDGEHALT_COL)
    stundensatz_col_idx = _col_letter_to_index(STUNDENSATZ_COL)

    mitarbeiter: List[MitarbeiterZeile] = []

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col_idx).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        info_val = ws.cell(r, info_col_idx).value
        info = str(info_val).strip() if info_val else None

        zeile = MitarbeiterZeile(name=name, pers_nr=None, info=info)

        tab = name_to_tab.get(name)
        if tab:
            zeile.pers_nr = _persnr_aus_tabname(tab)
        # Wenn kein Tab oder kein Suffix: pers_nr bleibt None, das wird über
        # die UI (übersprungen_keine_persnr / Bulk-Eingabe) behandelt — kein
        # Eintrag in ma.warnungen, weil das KEIN Datenfehler ist (z.B.
        # Festbezugs-Nebenjobber haben kein Stunden-Tab).

        zeile.stundensatz = _zahl(ws.cell(r, stundensatz_col_idx).value)

        for m, col_idx in mapped_cols:
            val = _zahl(ws.cell(r, col_idx).value)
            if val == 0:
                continue
            if m.get("umrechnen") == "eur_durch_stundensatz":
                if zeile.stundensatz <= 0:
                    zeile.warnungen.append(
                        f"{m['label']}: Umrechnung €→Std nicht möglich, "
                        f"Stundensatz fehlt oder ist 0 (Excel-Spalte {STUNDENSATZ_COL})."
                    )
                    continue
                val = val / zeile.stundensatz
            zeile.werte[m["lohnart"]] = round(val, 2)

        for u, col_idx in manuell_cols:
            val = _zahl(ws.cell(r, col_idx).value)
            if val != 0:
                zeile.manuell_werte[u["excel_header"]] = val

        zeile.soll_grundgehalt = _zahl(ws.cell(r, kontroll_col_idx).value)

        # Plausibilitätscheck: Soll-Grundgehalt sollte ~ Stunden × Stundensatz sein
        arbeitsstd = zeile.werte.get("1000", 0.0)
        if zeile.soll_grundgehalt and zeile.stundensatz and arbeitsstd:
            erwartet = arbeitsstd * zeile.stundensatz
            diff = abs(zeile.soll_grundgehalt - erwartet)
            if diff > 10.0:
                zeile.warnungen.append(
                    f"Plausibilität: Soll-Grundgehalt {zeile.soll_grundgehalt:.2f} € "
                    f"weicht von (Stunden {arbeitsstd:.2f} × Satz {zeile.stundensatz:.2f} = "
                    f"{erwartet:.2f} €) um {diff:.2f} € ab. Bitte Excel prüfen."
                )

        mitarbeiter.append(zeile)

    return ParseResult(mitarbeiter=mitarbeiter)


def monat_jahr_aus_dateiname(filename: str) -> Optional[Tuple[int, int]]:
    """'Gehaltsliste_2026_3 Wittys.xlsx' -> (2026, 3). None wenn nicht erkannt."""
    m = re.search(r"(\d{4})[_\-\s](\d{1,2})", filename)
    if not m:
        return None
    jahr, monat = int(m.group(1)), int(m.group(2))
    if 1 <= monat <= 12 and 2000 <= jahr <= 2100:
        return (jahr, monat)
    return None


def firma_aus_dateiname(filename: str) -> str:
    """'Gehaltsliste_2026_3 Wittys.xlsx' -> 'Wittys'.
    'Gehaltsliste 2026-05 Firma B.xlsx' -> 'Firma B'.
    Fallback: Dateiname ohne Extension."""
    base = re.sub(r"\.xlsx?$", "", filename, flags=re.IGNORECASE)
    m = re.search(r"\d{4}[_\-\s]\d{1,2}\s+(.+)$", base)
    if m:
        return m.group(1).strip()
    m = re.search(r"\d{4}[_\-\s]\d{1,2}[_\-\s]+(.+)$", base)
    if m:
        return m.group(1).strip()
    return base.strip()
